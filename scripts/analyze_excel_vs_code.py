import openpyxl
import re
import sys
from pathlib import Path

from runtime_recap_builder import AREA_SHEET_CANONICAL, find_area_sheet_name
from runtime_shift_rules import as_text, clean_nik, format_date, parse_sheet_date

# Force UTF-8 stdout encoding for Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

REPO_ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = REPO_ROOT / "EMPLOYEE DATA.xlsx"

# Standard headers from SharedLib.gs
EXPECTED_HEADERS = {
    'KARYAWAN': ['NIK', 'NAMA', 'TYPE KAYARAWAN', 'DEPT', 'JABATAN', 'USER LEVEL', 'PASSWORD'],
    'REGISTRASI SAAT MASUK PABRIK': ['NO KARTU MK', 'NIK', 'NAMA', 'TANGGAL', 'JAM MASUK', 'SHIFT'],
    'REGISTRASI SAAT KELUAR PABRIK': ['NO KARTU MK', 'NIK', 'NAMA', 'TANGGAL', 'JAM KELUAR', 'SHIFT'],
    'REGISTRASI MASUK KELUAR AREA KERJA': ['NO KARTU MK', 'INOUT', 'TANGGAL', 'JAM CATAT', 'NIK', 'NAMA', 'TUJUAN', 'CATATAN'],
    'BINDING_KARTU_MK': ['NO_KARTU_MK', 'NIK', 'NAMA', 'DEPT', 'JABATAN', 'WAKTU_BIND', 'STATUS'],
    'ABSEN IN OUT MK': ['TANGGAL', 'NIK', 'NAMA', 'DEPARTEMEN', 'JABATAN', 'JAM MASUK', 'JAM KELUAR', 'STATUS', 'NO KARTU MK', 'NO LOKER'],
    'JADWAL_SHIFT': ['NIK', 'NAMA', 'DEPT', 'SHIFT', 'TANGGAL_MULAI', 'TANGGAL_SELESAI']
}

def normalize_hdr(v):
    return re.sub(r'[\s_]+', '', str(v).strip().upper())


def normalize_key_date(value):
    parsed = parse_sheet_date(value)
    return format_date(parsed) if parsed else as_text(value).strip()

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    print("==================================================")
    print(f"ANALYSIS OF EXCEL: {EXCEL_PATH}")
    print(f"Tabs Found ({len(wb.sheetnames)}): {wb.sheetnames}")
    print("==================================================\n")

    # 1. Header Validation & Tab Name Matching
    print("--- 1. HEADER & TAB NAME VALIDATION ---")
    for expected_name, exp_headers in EXPECTED_HEADERS.items():
        actual_sheet_name = None
        if expected_name == AREA_SHEET_CANONICAL:
            actual_sheet_name = find_area_sheet_name(wb.sheetnames)
        else:
            for sname in wb.sheetnames:
                if sname == expected_name:
                    actual_sheet_name = sname
                    break

        if not actual_sheet_name:
            print(f"[MISSING] Sheet Missing in Excel: '{expected_name}'")
            continue
        
        if actual_sheet_name != expected_name:
            print(f"[TAB NAME MISMATCH] Code expects '{expected_name}', but Excel tab name is '{actual_sheet_name}'")

        ws = wb[actual_sheet_name]
        actual_headers = [as_text(cell.value) for cell in ws[1]]
        actual_norm = [normalize_hdr(h) for h in actual_headers if h]

        # Check required columns
        missing_cols = []
        for h in exp_headers:
            if normalize_hdr(h) not in actual_norm:
                missing_cols.append(h)
        
        if missing_cols:
            print(f"[HEADER MISMATCH] Sheet '{actual_sheet_name}': Missing Columns -> {missing_cols}")
            print(f"   Actual Headers: {actual_headers[:10]}")
        else:
            print(f"[OK] Sheet '{actual_sheet_name}': Headers match code specs ({len(actual_headers)} cols)")

    print("\n--- 2. DATA DEEP INSPECTION ---")
    
    # KARYAWAN Master Data
    if 'KARYAWAN' in wb.sheetnames:
        ws = wb['KARYAWAN']
        rows = list(ws.iter_rows(values_only=True))
        headers = [as_text(h) for h in rows[0]] if rows else []
        print(f"KARYAWAN Headers in Excel: {headers}")
        
        data_rows = rows[1:] if len(rows) > 1 else []
        niks = set()
        dup_niks = []
        missing_type = 0
        missing_dept = 0
        for idx, r in enumerate(data_rows, start=2):
            nik = clean_nik(r[0])
            if not nik: continue
            if nik in niks:
                dup_niks.append((nik, idx))
            niks.add(nik)
            if len(r) > 2 and not as_text(r[2]): missing_type += 1
            if len(r) > 3 and not as_text(r[3]): missing_dept += 1

        print(f"📋 KARYAWAN: Total Master = {len(niks)} NIKs | Duplicate NIKs = {len(dup_niks)} {dup_niks[:5]} | Missing Type = {missing_type} | Missing Dept = {missing_dept}")

    # BINDING_KARTU_MK
    if 'BINDING_KARTU_MK' in wb.sheetnames:
        ws = wb['BINDING_KARTU_MK']
        rows = list(ws.iter_rows(values_only=True))
        headers = [as_text(h) for h in rows[0]] if rows else []
        print(f"BINDING_KARTU_MK Headers in Excel: {headers}")
        
        data_rows = rows[1:] if len(rows) > 1 else []
        status_counts = {}
        bound_cards = {}
        bound_niks = {}
        dup_bound_cards = []
        dup_bound_niks = []

        for idx, r in enumerate(data_rows, start=2):
            card = as_text(r[0]).upper()
            nik = clean_nik(r[1])
            status = as_text(r[6]).upper() if len(r) > 6 else ''
            if not card and not nik: continue

            status_counts[status] = status_counts.get(status, 0) + 1

            if status == 'BOUND':
                if card in bound_cards:
                    dup_bound_cards.append((card, idx, bound_cards[card]))
                else:
                    bound_cards[card] = idx

                if nik in bound_niks:
                    dup_bound_niks.append((nik, idx, bound_niks[nik]))
                else:
                    bound_niks[nik] = idx

        print(f"📋 BINDING_KARTU_MK: Rows = {len(data_rows)} | Status Breakdown = {status_counts}")
        if dup_bound_cards:
            print(f"   [CRITICAL] Duplicate Active BOUND Cards = {len(dup_bound_cards)}: {dup_bound_cards[:5]}")
        if dup_bound_niks:
            print(f"   [CRITICAL] Duplicate Active BOUND NIKs = {len(dup_bound_niks)}: {dup_bound_niks[:5]}")

    # ABSEN IN OUT MK
    if 'ABSEN IN OUT MK' in wb.sheetnames:
        ws = wb['ABSEN IN OUT MK']
        rows = list(ws.iter_rows(values_only=True))
        headers = [as_text(h) for h in rows[0]] if rows else []
        print(f"ABSEN IN OUT MK Headers in Excel: {headers}")
        
        data_rows = rows[1:] if len(rows) > 1 else []
        status_counts = {}
        keys = {}
        duplicate_recap_keys = []
        di_dalam_rows = []
        keluar_tanpa_masuk_rows = []
        raw_date_types = set()

        for idx, r in enumerate(data_rows, start=2):
            raw_tgl = r[0]
            raw_date_types.add(type(raw_tgl).__name__)
            tgl = normalize_key_date(raw_tgl)
            nik = clean_nik(r[1])
            jam_masuk = as_text(r[5])
            jam_keluar = as_text(r[6])
            status = as_text(r[7]).upper() if len(r) > 7 else ''

            if not tgl and not nik: continue

            key = f"{tgl}|{nik}"
            if key in keys:
                duplicate_recap_keys.append((key, idx, keys[key]))
            else:
                keys[key] = idx

            status_counts[status] = status_counts.get(status, 0) + 1

            if status == 'DI DALAM' or (jam_masuk and not jam_keluar):
                di_dalam_rows.append((idx, tgl, nik, as_text(r[2]), jam_masuk))
            if status == 'KELUAR TANPA MASUK' or (jam_keluar and not jam_masuk):
                keluar_tanpa_masuk_rows.append((idx, tgl, nik, as_text(r[2]), jam_keluar))

        non_empty_rows = sum(1 for r in data_rows if any(cell not in (None, '') for cell in r))
        print(f"📋 ABSEN IN OUT MK: Physical Rows = {len(data_rows)} | Non-empty Rows = {non_empty_rows}")
        print(f"   Status Counts: {status_counts}")
        print(f"   Date Cell Types in Excel: {list(raw_date_types)}")
        print(f"   [ANOMALY] Duplicate Recap Keys (TANGGAL|NIK): {len(duplicate_recap_keys)}")
        if duplicate_recap_keys:
            print(f"      Examples (Key, CurrentRow, ExistingRow): {duplicate_recap_keys[:10]}")
        print(f"   [ANOMALY] Hanging 'DI DALAM' (Jam Keluar Kosong): {len(di_dalam_rows)}")
        if di_dalam_rows:
            print(f"      Recent 5: {di_dalam_rows[-5:]}")
        print(f"   [ANOMALY] Orphan 'KELUAR TANPA MASUK' (Jam Masuk Kosong): {len(keluar_tanpa_masuk_rows)}")
        if keluar_tanpa_masuk_rows:
            print(f"      Recent 5: {keluar_tanpa_masuk_rows[-5:]}")

    # REGISTRASI SAAT MASUK PABRIK
    if 'REGISTRASI SAAT MASUK PABRIK' in wb.sheetnames:
        ws = wb['REGISTRASI SAAT MASUK PABRIK']
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        headers = [as_text(h) for h in rows[0]] if rows else []
        print(f"📋 REGISTRASI SAAT MASUK PABRIK: Total Log Rows = {len(data_rows)} | Headers: {headers}")

    # REGISTRASI SAAT KELUAR PABRIK
    if 'REGISTRASI SAAT KELUAR PABRIK' in wb.sheetnames:
        ws = wb['REGISTRASI SAAT KELUAR PABRIK']
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        headers = [as_text(h) for h in rows[0]] if rows else []
        print(f"📋 REGISTRASI SAAT KELUAR PABRIK: Total Log Rows = {len(data_rows)} | Headers: {headers}")

    # REGISTRASI MASUK KELUAR AREA KERJA
    area_sheet = find_area_sheet_name(wb.sheetnames)
    if area_sheet:
        ws = wb[area_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        headers = [as_text(h) for h in rows[0]] if rows else []
        print(f"📋 AREA KERJA (Tab: '{area_sheet}'): Total Scan Rows = {len(data_rows)} | Headers: {headers}")

    # JADWAL_SHIFT
    if 'JADWAL_SHIFT' in wb.sheetnames:
        ws = wb['JADWAL_SHIFT']
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        headers = [as_text(h) for h in rows[0]] if rows else []
        print(f"📋 JADWAL_SHIFT: Total Scheduled Rows = {len(data_rows)} | Headers: {headers}")

if __name__ == '__main__':
    main()
