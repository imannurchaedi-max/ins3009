from __future__ import annotations

import json
import re
from collections import Counter
from datetime import date
from pathlib import Path
from typing import TypedDict

import openpyxl
from langgraph.graph import END, StateGraph

from runtime_recap_builder import (
    AREA_SHEET_CANONICAL,
    SHEET_KARYAWAN,
    SHEET_KELUAR,
    SHEET_MASUK,
    SHEET_RECAP,
    find_area_sheet_name,
    load_karyawan_map,
)
from runtime_shift_rules import as_text, clean_nik, format_date, parse_sheet_date


REPO_ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = REPO_ROOT / "EMPLOYEE DATA.xlsx"
GRAPHIFY_JSON = REPO_ROOT / "graphify-out" / "graph.json"
REPORT_PATH = REPO_ROOT / "reports" / "runtime_truth_audit_2026-08-01.md"
REPORT_JSON_PATH = REPO_ROOT / "reports" / "runtime_truth_audit_2026-08-01.json"
ACTIVE_ROOT = REPO_ROOT / "active" / "HOME_PORTAL"
CURRENT_DATE = date(2026, 8, 1)


class AuditState(TypedDict, total=False):
    workbook: dict
    runtime: dict
    graphify: dict
    report_path: str
    report_json_path: str


def profile_workbook(_: AuditState) -> AuditState:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    karyawan_map = load_karyawan_map(wb)
    recap_ws = wb[SHEET_RECAP]
    month_counts = Counter()
    status_counts = Counter()
    future_rows = 0
    duplicate_keys = 0
    seen_keys = set()

    for row in recap_ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        parsed = parse_sheet_date(row[0])
        tanggal = format_date(parsed) if parsed else as_text(row[0])
        nik = clean_nik(row[1])
        status_counts[as_text(row[7]).upper()] += 1
        key = f"{tanggal}|{nik}"
        if key in seen_keys:
            duplicate_keys += 1
        seen_keys.add(key)
        if parsed:
            month_counts[parsed.strftime("%Y-%m")] += 1
            if parsed > CURRENT_DATE:
                future_rows += 1

    binding_ws = wb["BINDING_KARTU_MK"]
    binding_status_counts = Counter()
    for row in binding_ws.iter_rows(min_row=2, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        binding_status_counts[as_text(row[6]).upper()] += 1

    area_sheet_name = find_area_sheet_name(wb.sheetnames)
    area_rows = 0
    if area_sheet_name:
        area_ws = wb[area_sheet_name]
        for row in area_ws.iter_rows(min_row=2, values_only=True):
            if any(cell not in (None, "") for cell in row):
                area_rows += 1

    return {
        "workbook": {
            "sheetCount": len(wb.sheetnames),
            "sheetNames": wb.sheetnames,
            "core": {
                "karyawanCount": len(karyawan_map),
                "masukRows": count_non_empty_rows(wb[SHEET_MASUK]),
                "keluarRows": count_non_empty_rows(wb[SHEET_KELUAR]),
                "recapRows": count_non_empty_rows(recap_ws),
                "areaRows": area_rows,
                "jadwalRows": count_non_empty_rows(wb["JADWAL_SHIFT"]),
            },
            "recap": {
                "statusCounts": dict(status_counts),
                "duplicateKeys": duplicate_keys,
                "monthCounts": dict(sorted(month_counts.items())),
                "futureRowsAfter_2026_08_01": future_rows,
            },
            "binding": {
                "statusCounts": dict(binding_status_counts),
            },
            "areaSheetName": area_sheet_name,
        }
    }


def inspect_runtime(state: AuditState) -> AuditState:
    gate_text = (ACTIVE_ROOT / "GateFunctions.gs").read_text(encoding="utf-8")
    area_text = (ACTIVE_ROOT / "AreaFunctions.gs").read_text(encoding="utf-8")
    report_text = (ACTIVE_ROOT / "ReportFunctions.gs").read_text(encoding="utf-8")
    repair_text = (ACTIVE_ROOT / "DataRepairUtils.gs").read_text(encoding="utf-8")
    shared_text = (ACTIVE_ROOT / "SharedLib.gs").read_text(encoding="utf-8")
    app_text = (ACTIVE_ROOT / "app.html").read_text(encoding="utf-8")

    frontend_calls = sorted(set(re.findall(r"\.([A-Za-z0-9_]+)\(", app_text)))
    server_funcs = sorted(set(re.findall(r"function\s+([A-Za-z0-9_]+)\s*\(", gate_text + area_text + report_text + repair_text + shared_text)))

    runtime = {
        "frontendGoogleScriptRunCalls": [name for name in frontend_calls if name in server_funcs],
        "criticalFlows": [
            {
                "name": "Masuk Pabrik",
                "frontend": "confirmMasuk() -> google.script.run.bindKartu()",
                "backend": [
                    "bindKartu()",
                    "getKaryawanByNIK()",
                    "getBindingStatus()",
                    "detectShift()",
                    "safeUpdateRecapAbsen()",
                    "rebuildHistoricalRecapDataset_()",
                ],
                "sheetsRead": ["KARYAWAN", "BINDING_KARTU_MK"],
                "sheetsWrite": ["REGISTRASI SAAT MASUK PABRIK", "BINDING_KARTU_MK", "ABSEN IN OUT MK"],
            },
            {
                "name": "Keluar Pabrik",
                "frontend": "confirmKeluar() -> google.script.run.releaseKartu()",
                "backend": [
                    "releaseKartu()",
                    "getBindingStatus()",
                    "resolveFactoryWorkDate()",
                    "detectShift()",
                    "safeUpdateRecapAbsen()",
                    "rebuildHistoricalRecapDataset_()",
                ],
                "sheetsRead": ["BINDING_KARTU_MK", "KARYAWAN"],
                "sheetsWrite": ["REGISTRASI SAAT KELUAR PABRIK", "BINDING_KARTU_MK", "ABSEN IN OUT MK"],
            },
            {
                "name": "Scan Area Kerja",
                "frontend": "handleSecurityScan()/onSerialScanned() -> google.script.run.scanAreaKerja()",
                "backend": [
                    "scanAreaKerja()",
                    "getBindingStatus()",
                    "getFactoryFlowStatusFromLogs_()",
                ],
                "sheetsRead": ["BINDING_KARTU_MK", "ABSEN IN OUT MK", "KARYAWAN"],
                "sheetsWrite": ["REGISTRASI MASUK KELUAR AREA KERJA"],
            },
            {
                "name": "Laporan Absen",
                "frontend": "processAbsenReport() -> google.script.run.getAbsenReport()",
                "backend": [
                    "getAbsenReport()",
                    "getAbsenReportFullData_()",
                    "buildPaginationMeta_()",
                ],
                "sheetsRead": ["ABSEN IN OUT MK"],
                "sheetsWrite": [],
            },
            {
                "name": "Laporan Area",
                "frontend": "processAreaReport() -> google.script.run.getAreaActivityReport()",
                "backend": [
                    "getAreaActivityReport()",
                    "getAreaActivityReportFullData_()",
                    "buildPaginationMeta_()",
                ],
                "sheetsRead": ["REGISTRASI MASUK KELUAR AREA KERJA", "KARYAWAN"],
                "sheetsWrite": [],
            },
            {
                "name": "Perbaikan Spreadsheet",
                "frontend": "Spreadsheet menu -> fixAllSpreadsheetErrors()",
                "backend": [
                    "fixAllSpreadsheetErrors()",
                    "repairFactoryMasukLog_()",
                    "repairFactoryKeluarLog_()",
                    "rebuildHistoricalRecapDataset_()",
                    "buildFactoryRecapRowsFromEvents_()",
                ],
                "sheetsRead": ["REGISTRASI SAAT MASUK PABRIK", "REGISTRASI SAAT KELUAR PABRIK", "KARYAWAN", "ABSEN IN OUT MK", "BINDING_KARTU_MK", "JADWAL_SHIFT"],
                "sheetsWrite": ["ABSEN IN OUT MK", "BINDING_KARTU_MK", "REGISTRASI SAAT MASUK PABRIK", "REGISTRASI SAAT KELUAR PABRIK"],
            },
        ],
        "runtimeNotes": [
            "Source of truth write-path ada di active/HOME_PORTAL/.",
            "ABSEN IN OUT MK diperlakukan sebagai recap turunan, bukan log primer.",
            "JADWAL_SHIFT saat ini kosong, jadi expected shift override belum aktif.",
            "Pagination server-side di report aktif dengan DEFAULT_REPORT_PAGE_SIZE=25 dan MAX_REPORT_PAGE_SIZE=100.",
        ],
    }
    return {"runtime": runtime}


def inspect_graphify(_: AuditState) -> AuditState:
    summary = {
        "available": GRAPHIFY_JSON.exists(),
        "homePortalNodeCount": 0,
        "homePortalFiles": [],
        "keySymbolsPresent": {},
    }
    if not GRAPHIFY_JSON.exists():
        return {"graphify": summary}

    data = json.loads(GRAPHIFY_JSON.read_text(encoding="utf-8"))
    nodes = data.get("nodes", [])
    home_nodes = [node for node in nodes if str(node.get("source_file", "")).startswith("active/HOME_PORTAL/")]
    summary["homePortalNodeCount"] = len(home_nodes)
    summary["homePortalFiles"] = sorted(set(node.get("source_file") for node in home_nodes if node.get("source_file")))
    key_labels = ["bindKartu()", "releaseKartu()", "scanAreaKerja()", "getAbsenReport()", "getAreaActivityReport()", "fixAllSpreadsheetErrors()"]
    labels = {node.get("label"): node for node in home_nodes}
    summary["keySymbolsPresent"] = {label: label in labels for label in key_labels}
    return {"graphify": summary}


def write_report(state: AuditState) -> AuditState:
    workbook = state["workbook"]
    runtime = state["runtime"]
    graphify = state["graphify"]

    lines = []
    lines.append("# Runtime Truth Audit - 2026-08-01")
    lines.append("")
    lines.append("## Ringkasan Eksekutif")
    lines.append("")
    lines.append("- Workbook lokal `EMPLOYEE DATA.xlsx` sekarang bisa dibaca konsisten tab-per-tab dan recap sudah dinormalkan ke string tanggal serta sheet area kerja canonical.")
    lines.append("- Flow yang benar adalah satu arah: log masuk/keluar pabrik sebagai sumber primer, log area sebagai sumber primer area, lalu `ABSEN IN OUT MK` hanya recap turunan.")
    lines.append("- Masalah utama bukan seluruh aturan shift, melainkan recap historis yang sebelumnya tidak sinkron dengan log primer, format tanggal campur, nama sheet area terpotong, dan alat audit lama yang memakai aturan shift usang.")
    lines.append("- `JADWAL_SHIFT` kosong, jadi semua logika expected shift berbasis jadwal saat ini praktis tidak berperan.")
    lines.append("")
    lines.append("## Fakta Workbook")
    lines.append("")
    lines.append(f"- Total sheet: {workbook['sheetCount']}")
    lines.append(f"- Sheet area kerja aktual: `{workbook['areaSheetName']}`")
    lines.append(f"- Master karyawan: {workbook['core']['karyawanCount']} NIK")
    lines.append(f"- Log masuk pabrik: {workbook['core']['masukRows']} baris")
    lines.append(f"- Log keluar pabrik: {workbook['core']['keluarRows']} baris")
    lines.append(f"- Recap absen: {workbook['core']['recapRows']} baris")
    lines.append(f"- Log area kerja: {workbook['core']['areaRows']} baris")
    lines.append(f"- Jadwal shift: {workbook['core']['jadwalRows']} baris")
    lines.append(f"- Status binding: {workbook['binding']['statusCounts']}")
    lines.append(f"- Status recap: {workbook['recap']['statusCounts']}")
    lines.append(f"- Duplicate key recap `(tanggal|nik)`: {workbook['recap']['duplicateKeys']}")
    lines.append(f"- Baris recap bertanggal setelah 2026-08-01: {workbook['recap']['futureRowsAfter_2026_08_01']}")
    lines.append("")
    lines.append("### Sebaran Recap per Bulan")
    lines.append("")
    for month, count in workbook["recap"]["monthCounts"].items():
        lines.append(f"- {month}: {count} baris")
    lines.append("")
    lines.append("## One-Direction Multipath Flow")
    lines.append("")
    lines.append("1. Karyawan datang ke pabrik -> `bindKartu()` memvalidasi NIK dan kartu, menulis log ke `REGISTRASI SAAT MASUK PABRIK`, lalu memperbarui state binding.")
    lines.append("2. Setelah log masuk tercatat, recap `ABSEN IN OUT MK` dibangun dari log primer, bukan dari input manual terpisah.")
    lines.append("3. Selama di area kerja -> `scanAreaKerja()` hanya menulis log IN/OUT area ke `REGISTRASI MASUK KELUAR AREA KERJA` berdasarkan status binding dan status pabrik dari log.")
    lines.append("4. Karyawan keluar pabrik -> `releaseKartu()` menulis log ke `REGISTRASI SAAT KELUAR PABRIK`, melepas binding, lalu rebuild recap lagi.")
    lines.append("5. Dashboard dan report hanya membaca data turunan atau read model: recap pabrik, log area, master karyawan, dan jadwal jika suatu saat diisi.")
    lines.append("")
    lines.append("## Dependency Map Kritis")
    lines.append("")
    for flow in runtime["criticalFlows"]:
        lines.append(f"### {flow['name']}")
        lines.append(f"- Frontend caller: {flow['frontend']}")
        lines.append(f"- Backend chain: {', '.join(flow['backend'])}")
        lines.append(f"- Read sheet: {', '.join(flow['sheetsRead']) if flow['sheetsRead'] else '-'}")
        lines.append(f"- Write sheet: {', '.join(flow['sheetsWrite']) if flow['sheetsWrite'] else '-'}")
        lines.append("")
    lines.append("## Graphify dan GitNexus")
    lines.append("")
    if graphify["available"]:
        lines.append(f"- Graphify artifact tersedia di `graphify-out/graph.json` dengan {graphify['homePortalNodeCount']} node yang berasal dari `active/HOME_PORTAL/`.")
        lines.append(f"- File HOME_PORTAL yang terlihat di graphify: {', '.join(graphify['homePortalFiles'])}")
        lines.append(f"- Simbol penting terdeteksi: {graphify['keySymbolsPresent']}")
    else:
        lines.append("- Graphify artifact tidak ditemukan.")
    lines.append("- GitNexus graph tersedia dan bisa dipakai untuk context/doctor, tetapi FTS index di mesin ini masih degradasi sehingga query teks bebas tidak selalu bisa dipercaya.")
    lines.append("")
    lines.append("## Temuan Arsitektur")
    lines.append("")
    lines.append("- `ABSEN IN OUT MK` tidak boleh lagi menjadi sumber keputusan operasional primer. Ia harus diperlakukan sebagai materialized view yang selalu bisa dibangun ulang dari log masuk dan log keluar.")
    lines.append("- `REGISTRASI MASUK KELUAR AREA KERJA` tidak boleh menentukan status pabrik. Ia hanya merekam pergerakan area setelah status di pabrik valid.")
    lines.append("- `JADWAL_SHIFT` saat ini kosong, jadi semua evaluasi shift masih murni berdasar jam scan dan `SHIFT_CONFIG` runtime.")
    lines.append("- Workbook mengandung data masa depan terhadap tanggal audit 2026-08-01. Ini perlu dipastikan apakah memang data operasional yang sah atau hasil input/ekspor lintas periode.")
    lines.append("")
    lines.append("## Urutan Perbaikan yang Disarankan")
    lines.append("")
    lines.append("1. Kunci model data primer: log masuk, log keluar, binding, area log, master karyawan.")
    lines.append("2. Satukan semua alat audit/perbaikan ke helper aturan runtime aktif.")
    lines.append("3. Ubah backend report supaya semua tabel besar memakai pagination server-side dan tidak mengirim payload besar ke HTML.")
    lines.append("4. Audit semua fungsi yang masih membaca recap lama sebagai sumber status, lalu ganti ke pembacaan dari log primer atau hasil rebuild yang baru.")
    lines.append("5. Setelah runtime stabil, baru rapikan UX dashboard per role: Security Personel, Area Owner, HR Supervisor, HR Manager.")
    lines.append("")

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")
    REPORT_JSON_PATH.write_text(
        json.dumps(
            {
                "generatedAt": "2026-08-01",
                "workbook": workbook,
                "runtime": runtime,
                "graphify": graphify,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return {
        "report_path": str(REPORT_PATH),
        "report_json_path": str(REPORT_JSON_PATH),
    }


def count_non_empty_rows(ws):
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell not in (None, "") for cell in row):
            count += 1
    return count


def main():
    graph = StateGraph(AuditState)
    graph.add_node("profile_workbook", profile_workbook)
    graph.add_node("inspect_runtime", inspect_runtime)
    graph.add_node("inspect_graphify", inspect_graphify)
    graph.add_node("write_report", write_report)
    graph.set_entry_point("profile_workbook")
    graph.add_edge("profile_workbook", "inspect_runtime")
    graph.add_edge("inspect_runtime", "inspect_graphify")
    graph.add_edge("inspect_graphify", "write_report")
    graph.add_edge("write_report", END)

    app = graph.compile()
    result = app.invoke({})
    print(f"Markdown report: {result['report_path']}")
    print(f"JSON report    : {result['report_json_path']}")


if __name__ == "__main__":
    main()
