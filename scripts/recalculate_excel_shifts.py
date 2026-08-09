import openpyxl
import sys
from pathlib import Path

from runtime_recap_builder import SHEET_KELUAR, SHEET_MASUK, repair_shift_labels


if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")


REPO_ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = REPO_ROOT / "EMPLOYEE DATA.xlsx"


def print_samples(title, samples):
    print(title)
    if not samples:
        print("  Tidak ada perbedaan shift.")
        return
    for sample in samples:
        print(
            "  Row {row:4d} | Jam: {time:8s} | Old: {oldShift:8s} -> New: {newShift:8s} | NIK: {nik}".format(
                **sample
            )
        )


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    masuk_result = repair_shift_labels(wb[SHEET_MASUK], "masuk")
    keluar_result = repair_shift_labels(wb[SHEET_KELUAR], "keluar")

    print("--- AUDIT SHIFT REGISTRASI SAAT MASUK PABRIK ---")
    print_samples("Contoh koreksi masuk:", masuk_result["samples"])
    print(
        f"Total log masuk yang berbeda dari aturan runtime aktif: {masuk_result['fixed']} / {wb[SHEET_MASUK].max_row - 1}"
    )

    print("\n--- AUDIT SHIFT REGISTRASI SAAT KELUAR PABRIK ---")
    print_samples("Contoh koreksi keluar:", keluar_result["samples"])
    print(
        f"Total log keluar yang berbeda dari aturan runtime aktif: {keluar_result['fixed']} / {wb[SHEET_KELUAR].max_row - 1}"
    )


if __name__ == "__main__":
    main()
